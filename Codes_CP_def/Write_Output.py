from openpyxl import Workbook
import openpyxl
import os
from pyomo.environ import *

def write(excel_path, Param=None, model=None, sol=None, step=None, exit=None, time=None):

    if step =='Param':
        workbook = Workbook()

        workbook.remove(workbook[workbook.sheetnames[0]])

        worksheet = workbook.create_sheet("Param")

        worksheet.append(["Param", "Value"])

        for i in Param:
            worksheet.append([i, Param[i]])
        
        workbook.save(excel_path)


    elif step =='Global':
        workbook = openpyxl.load_workbook(excel_path)

        worksheet = workbook.create_sheet("PPC_var")

        worksheet.append(["Terminal condition", exit, "Computational Time", time])

        if exit == 'Feasible':
            worksheet.append(["Variable", "Index",	"Value"])
            for v in model.get_all_variables():
                name=v.name.split(':')
                worksheet.append([name[0], name[1],	str(sol.get_value(v))])
    
        workbook.save(excel_path)