from openpyxl import Workbook
import openpyxl
import os
from pyomo.environ import *

def write(excel_path, model, step, num_chg= None, exit=None, time=None):

    if step =='Param':
        workbook = Workbook()

        workbook.remove(workbook[workbook.sheetnames[0]])

        worksheet = workbook.create_sheet("Param")

        worksheet.append(["Param", "Value"])

        for i in model.component_data_objects(Param, active=True):
            worksheet.append([i.name, value(i)])


    elif step =='Global':
        workbook = openpyxl.load_workbook(excel_path)

        worksheet = workbook.create_sheet("Global")

        worksheet.append(["Terminal condition", exit, "Computational Time", time])

        worksheet.append(["Variable", "Index",	"Value"])

        for i in model.component_data_objects(Var, active=True):
            if 'eps' not in i.name:
                name=i.name.replace('[', ':').replace(']', '').split(':')
                if exit not in ['infeasible', 'Time limit']:
                    worksheet.append([name[0], name[1],	value(i)])
                else:
                    worksheet.append([name[0], name[1], 'None'])
    
    elif step == '1':

        workbook = openpyxl.load_workbook(excel_path)

        worksheet = workbook.create_sheet("step_1")

        worksheet.append(["Terminal condition", exit, "Computational Time", time])

        worksheet.append(["Variable", "Index",	"Value"])

        for i in model.component_data_objects(Var, active=True):
            if 'u' not in i.name and 'eps' not in i.name:
                name=i.name.replace('[', ':').replace(']', '').split(':')
                if exit not in ['infeasible', 'Time limit']:
                    worksheet.append([name[0], name[1],	value(i)])
                else:
                    worksheet.append([name[0], name[1], 'None'])
    
    elif step == '2':

        workbook = openpyxl.load_workbook(excel_path)
        
        # if len(workbook.sheetnames) >=2:
        #     workbook.remove(workbook[workbook.sheetnames[1]])

        worksheet = workbook.create_sheet("step_3")

        worksheet.append(["Terminal condition", exit, "Computational Time", time])

        worksheet.append(["Variable", "Index",	"Value"])

        for i in model.component_data_objects(Var, active=True):
            if 'eps' in i.name or 'delta' in i.name:
                name=i.name.replace('[', ':').replace(']', '').split(':')
                if exit not in ['infeasible', 'Time limit']:
                    worksheet.append([name[0], name[1],	value(i)])
                else:
                    worksheet.append([name[0], name[1], 'None'])
    
    elif step == '3':
    
        workbook = openpyxl.load_workbook(excel_path)
        
        # if len(workbook.sheetnames) >=2:
        #     workbook.remove(workbook[workbook.sheetnames[1]])

        worksheet = workbook.create_sheet("step_2_num_chg_"+str(num_chg))

        worksheet.append(["Terminal condition", exit, "Computational Time", time])

        worksheet.append(["Variable", "Index",	"Value"])

        for i in model.component_data_objects(Var, active=True):
            if 'u' in i.name or 'a' in i.name or 't' in i.name or 'y' in i.name or '_q' in i.name or 'q_' in i.name:
                name=i.name.replace('[', ':').replace(']', '').split(':')
                if exit not in ['infeasible', 'Time limit']:
                    worksheet.append([name[0], name[1],	value(i)])
                else:
                    worksheet.append([name[0], name[1], 'None'])

    workbook.save(excel_path)